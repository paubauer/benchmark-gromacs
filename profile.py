import os
import csv
import argparse

import openpyxl
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, Border, Side, Color, Alignment
from openpyxl.formatting.rule import ColorScale, FormatObject
from openpyxl.utils import get_column_letter

# Usage:
# python3 profile.py --platform=CUDA --output=output.xlsx --title=titanv
# python3 profile.py --platform=HIP --input=output.xlsx --output=output1.xlsx --compare=titanv --title=2021-08-05

def timeMultiplier(s):
    if s == 's':
        return 1e6
    if s == 'ms':
        return 1e3
    if s == 'us':
        return 1
    raise Exception(f'Unknown time measure {s}')

def applyRule(ws, cell, compareCol):
    compareCell = ws.cell(row=cell.row, column=compareCol)
    if compareCell.value is not None:
        v = f'{get_column_letter(compareCell.column)}{compareCell.row}'
        first = FormatObject(type='formula', val=f'$A$1*{v}')
        mid = FormatObject(type='formula', val=f'{v}')
        last = FormatObject(type='formula', val=f'$A$2*{v}')
        colors = [Color('84ee00'), Color('ffff00'), Color('ff0000')]
        cs = ColorScale(cfvo=[first, mid, last], color=colors)
        rule = Rule(type='colorScale', colorScale=cs)
        ws.conditional_formatting.add(cell.coordinate, rule)

precisions = [
    'single',
    'mixed',
    'double',
]
tests = [
    ('adh_dodec_stream_commands', 10000),
    ('adh_dodec_mi100_commands', 10000),
    ('adh_dodec_mi200_commands', 10000),
    ('stmv_stream_commands', 10000),
    ('stmv_dodec_mi100_commands', 10000),
    ('stmv_dodec_mi200_commands', 10000),
    ('celluloze_nve_stream_commands', 10000),
    ('celluloze_nve_dodec_mi100_commands', 10000),
    ('celluloze_nve_dodec_mi200_commands', 10000)
]

parser = argparse.ArgumentParser()
parser.add_argument('--platform', default='HIP', dest='platform', choices=['CUDA', 'HIP'], help='name of the platform to profile')
parser.add_argument('--precision', default=None, dest='precision', choices=precisions, help='precision mode: single, mixed, double')
parser.add_argument('--test', default=None, dest='test', choices=[t[0] for t in tests])
parser.add_argument('--title', required=True, dest='title', help='title of a new column')
parser.add_argument('--compare', default=None, dest='compare', help='title of a column to compare with conditional formatting')
parser.add_argument('--input', default=None, dest='input')
parser.add_argument('--output', default='output.xlsx', dest='output')
args = parser.parse_args()
if args.precision:
    precisions = [args.precision]
if args.test:
    tests = [t for t in tests if t[0] == args.test]

wb = None
if not args.input:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
else:
    wb = openpyxl.load_workbook(args.input)

borderLeft = Border(left=Side(border_style='thin', color='000000'))
borderRight = Border(right=Side(border_style='thin', color='000000'))

for precision in precisions:
    for (test, steps) in tests:
        results = []
        resultsFileName = 'results.stats.csv'
        if args.platform == 'HIP':
            r = os.system(f'rocprof --basenames on --stats python3 benchmark.py --test={test} --steps={steps} --profile')
            if r:
                continue
            # resultsFileName = '2021-07-19-gbsa-mi100-0.csv'
            with open(resultsFileName, newline='') as csvfile:
                reader = csv.reader(csvfile)
                next(reader)
                for row in reader:
                    results.append((
                        row[0][:-3],
                        int(row[1]),
                        float(row[2]) / 1e3,
                        float(row[3]) / 1e3,
                        float(row[4]) / 100
                    ))
        elif args.platform == 'CUDA':
            r = os.system(f'nvprof --demangling off --profile-api-trace none --concurrent-kernels off --csv --log-file {resultsFileName} python3 benchmark.py --test={test} --steps={steps} --profile')
            if r:
                continue
            # resultsFileName = '2021-07-19-gbsa-titanv-0.csv'
            with open(resultsFileName, newline='') as csvfile:
                reader = csv.reader(csvfile)
                # nvprof adds some info before the table
                while next(reader)[0] != 'Type':
                    pass
                m = next(reader)
                m2 = timeMultiplier(m[2])
                m4 = timeMultiplier(m[4])
                for row in reader:
                    # nvprof adds some info after the table
                    if len(row) < 8:
                        break
                    results.append((
                        row[7],
                        int(row[3]),
                        float(row[2]) * m2,
                        float(row[4]) * m4,
                        float(row[1]) / 100
                    ))
        wsName = f'{precision[0]}-{test}'
        ws = None
        compareCol = None
        cs = 0
        isOk = True
        if not args.input:
            ws = wb.create_sheet(wsName)
            ri = 3
            for r in results:
                    ws.cell(row=ri, column=1, value=r[0])
                    ri += 1
            ws.column_dimensions['A'].width = 30
            ws.freeze_panes = 'B3'
            cs = 2
        else:
            ws = wb[wsName]
            cs = ws.max_column + 1
            if args.compare:
                for col in ws.iter_cols(min_row=1, max_row=1):
                    if col[0].value == args.compare:
                        compareCol = col[0].column
                        break
            if not compareCol:
                print(f'Column with value {args.compare} not found on {wsName}')
            for col in ws.iter_cols(min_row=1, max_row=1):
                if col[0].value == args.title:
                    print(f'Column with value {args.title} already exists on {wsName}')
                    isOk = False
                    break
        if isOk:
            # Used as low and high values for conditional formatting
            ws['A1'].value = 0.5
            ws['A2'].value = 3.0
            for r in results:
                ri = 3
                # Find a kernel with by name or add a new row
                while True:
                    c = ws.cell(row=ri, column=1)
                    if not c.value:
                        c.value = r[0]
                        break
                    if c.value == r[0]:
                        break
                    ri += 1
                c1 = ws.cell(row=ri, column=cs + 0, value=r[1])
                c1.number_format = '#,##0'
                c1.border = borderLeft
                c2 = ws.cell(row=ri, column=cs + 1, value=r[2])
                c2.number_format = '#,##0' # numbers.FORMAT_NUMBER_00
                if compareCol:
                    applyRule(ws, c2, compareCol + 1)
                c3 = ws.cell(row=ri, column=cs + 2, value=r[3])
                c3.number_format = '#,##0.0' # numbers.FORMAT_NUMBER_00
                if compareCol:
                    applyRule(ws, c3, compareCol + 2)
                c4 = ws.cell(row=ri, column=cs + 3, value=r[4])
                c4.number_format = '0.0%' # numbers.FORMAT_PERCENTAGE_00
                c4.border = borderRight
            ws.cell(row=1, column=cs + 0, value=args.title).alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=1, start_column=cs + 0, end_row=1, end_column=cs + 3)
            if compareCol:
                ws.cell(row=2, column=cs + 0, value='vs ' + args.compare).alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=2, start_column=cs + 0, end_row=2, end_column=cs + 3)
            ws.column_dimensions[get_column_letter(cs + 0)].border = borderLeft
            ws.column_dimensions[get_column_letter(cs + 3)].border = borderRight
            ws.column_dimensions[get_column_letter(cs + 0)].width = 8
            ws.column_dimensions[get_column_letter(cs + 1)].width = 12
            ws.column_dimensions[get_column_letter(cs + 2)].width = 10
            ws.column_dimensions[get_column_letter(cs + 3)].width = 8
wb.save(args.output)
